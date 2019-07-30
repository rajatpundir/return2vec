import os, requests, pandas as pd, openpyxl, datetime
from openpyxl.utils.dataframe import dataframe_to_rows
from itertools import islice

class Stock:
    def __init__(self, ticker, api_key, update_intraday=False, update_daily=False, interval = 1):
        self.api_key = api_key
        self.ticker = ticker
        if update_daily is not None:
            self.df_daily = None
            if update_daily:
                self.update_daily()
            else:
                self.load_dataframe_daily()
                if len(self.df_daily.index) != 0:
                    print('[' + self.ticker + ']', 'Daily :', self.df_daily.day[-1])
                    print('[' + self.ticker + ']', len(self.df_daily.index), 'values loaded.')
        if update_intraday is not None:
            self.df = None
            self.interval = interval
            if update_intraday:
                self.update_intraday()
            else:
                self.load_dataframe_intraday()
                if self.ticker.startswith('NSE') or self.ticker.startswith('BSE'):
                    self.df.date = self.df.date.dt.tz_localize(tz='US/Eastern').dt.tz_convert('Asia/Kolkata')
                if len(self.df.index) != 0:
                    print('[' + self.ticker + ']', 'Intraday :', str(self.df.date[-1:]).split('\n')[0].split(' ')[-2] + ' ' + str(self.df.date[-1:]).split('\n')[0].split(' ')[-1][:8])
                    print('[' + self.ticker + ']', len(self.df.index), 'values loaded.')
        
    def load_dataframe_intraday(self):
        if not os.path.exists('data'):
            os.mkdir('data')
        filename = './data/' + self.ticker + '.xlsx'
        if not os.path.exists(filename):
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(['date', 'open', 'high', 'low', 'close', 'volume'])
            wb.save(filename)
            wb.close()
        wb = openpyxl.load_workbook(filename)
        ws = wb.active
        data = ws.values
        wb.close()
        cols = next(data)[1:]
        cols = [c.lower() for c in cols]
        data = list(data)
        idx = [r[0] for r in data]
        data = (islice(r, 1, None) for r in data)
        self.df = pd.DataFrame(data, index=idx, columns=cols).reset_index()
        self.df.rename(index=str, columns={"index": "date"}, inplace=True)
        self.df.date = pd.to_datetime(self.df.date, format='%Y-%m-%d %H:%M:%S', exact=False)
        self.df.open = pd.to_numeric(self.df.open)
        self.df.high = pd.to_numeric(self.df.high)
        self.df.low = pd.to_numeric(self.df.low)
        self.df.close = pd.to_numeric(self.df.close)
        self.df.volume = pd.to_numeric(self.df.volume)
        self.df.fillna(0, inplace=True)
        self.df.sort_values(by=['date'], kind='quicksort', ascending=True, inplace=True)
        self.df.drop_duplicates(subset='date', keep='first', inplace=True)
        
    def load_dataframe_daily(self):
        if not os.path.exists('data'):
            os.mkdir('data')
        filename = './data/' + self.ticker + '_DAILY.xlsx'
        if not os.path.exists(filename):
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(['day', 'open', 'high', 'low', 'close', 'adjusted_close', 'volume', 'divident_amount', 'split_coefficient'])
            wb.save(filename)
            wb.close()
        wb = openpyxl.load_workbook(filename)
        ws = wb.active
        data = ws.values
        wb.close()
        cols = next(data)[1:]
        cols = [c.lower() for c in cols]
        data = list(data)
        idx = [r[0] for r in data]
        data = (islice(r, 1, None) for r in data)
        self.df_daily = pd.DataFrame(data, index=idx, columns=cols).reset_index()
        self.df_daily.rename(index=str, columns={"index": "day"}, inplace=True)
        self.df_daily.day = pd.to_datetime(self.df_daily.day, format='%Y-%m-%d', exact=False)
        self.df_daily.open = pd.to_numeric(self.df_daily.open)
        self.df_daily.high = pd.to_numeric(self.df_daily.high)
        self.df_daily.low = pd.to_numeric(self.df_daily.low)
        self.df_daily.close = pd.to_numeric(self.df_daily.close)
        self.df_daily.adjusted_close = pd.to_numeric(self.df_daily.adjusted_close)
        self.df_daily.volume = pd.to_numeric(self.df_daily.volume)
        self.df_daily.divident_amount = pd.to_numeric(self.df_daily.divident_amount)
        self.df_daily.split_coefficient = pd.to_numeric(self.df_daily.split_coefficient)
        self.df_daily.fillna(0, inplace=True)
        self.df_daily.fillna(0, inplace=True)
        self.df_daily.sort_values(by=['day'], kind='quicksort', ascending=True, inplace=True)
        self.df_daily.drop_duplicates(subset='day', keep='first', inplace=True)
            
    def save_dataframe_intraday(self):
        if not os.path.exists('data'):
            os.mkdir('data')
        filename = './data/' + self.ticker + '.xlsx'
        wb = openpyxl.Workbook()
        ws = wb.active
        for r in dataframe_to_rows(self.df, index=False, header=True):
            ws.append(r)
        wb.save(filename)
        wb.close()
        
    def save_dataframe_daily(self):
        if not os.path.exists('data'):
            os.mkdir('data')
        filename = './data/' + self.ticker + '_DAILY.xlsx'
        wb = openpyxl.Workbook()
        ws = wb.active
        for r in dataframe_to_rows(self.df_daily, index=False, header=True):
            ws.append(r)
        wb.save(filename)
        wb.close()
            
    def update_daily(self):
        self.load_dataframe_daily()
        outputsize = 'full'
        if len(self.df_daily.index) > 100:
            outputsize = 'compact'
        req = requests.get('https://www.alphavantage.co/query?function=TIME_SERIES_DAILY_ADJUSTED&symbol=' + self.ticker + '&outputsize=' + outputsize + '&apikey=' + self.api_key()).json()
        self.ticker = req['Meta Data']['2. Symbol']
        print('[' + self.ticker + ']', 'Daily :', req['Meta Data']['3. Last Refreshed'])
        before = len(self.df_daily.index)
        # Load new values
        new_df_daily = pd.DataFrame.from_dict(req[[k for k, v in req.items()][-1]], orient="index").reset_index()
        new_df_daily.columns = ['day', 'open', 'high', 'low', 'close', 'adjusted_close', 'volume', 'divident_amount', 'split_coefficient']
        new_df_daily.day = pd.to_datetime(new_df_daily.day, format='%Y-%m-%d', exact=False)
        new_df_daily.open = pd.to_numeric(new_df_daily.open)
        new_df_daily.high = pd.to_numeric(new_df_daily.high)
        new_df_daily.low = pd.to_numeric(new_df_daily.low)
        new_df_daily.close = pd.to_numeric(new_df_daily.close)
        new_df_daily.adjusted_close = pd.to_numeric(new_df_daily.adjusted_close)
        new_df_daily.volume = pd.to_numeric(new_df_daily.volume)
        new_df_daily.divident_amount = pd.to_numeric(new_df_daily.divident_amount)
        new_df_daily.split_coefficient = pd.to_numeric(new_df_daily.split_coefficient)
        new_df_daily.fillna(0, inplace=True)
        new_df_daily.sort_values(by=['day'], kind='quicksort', ascending=True, inplace=True)
        # Merge dataframes
        self.df_daily = new_df_daily.append(self.df_daily, ignore_index=True)
        self.df_daily.day = pd.to_datetime(self.df_daily.day, format='%Y-%m-%d', exact=False)
        self.df_daily.open = pd.to_numeric(self.df_daily.open)
        self.df_daily.high = pd.to_numeric(self.df_daily.high)
        self.df_daily.low = pd.to_numeric(self.df_daily.low)
        self.df_daily.close = pd.to_numeric(self.df_daily.close)
        self.df_daily.adjusted_close = pd.to_numeric(self.df_daily.adjusted_close)
        self.df_daily.volume = pd.to_numeric(self.df_daily.volume)
        self.df_daily.divident_amount = pd.to_numeric(self.df_daily.divident_amount)
        self.df_daily.split_coefficient = pd.to_numeric(self.df_daily.split_coefficient)
        self.df_daily.fillna(0, inplace=True)
        self.df_daily.sort_values(by=['day'], kind='mergesort', ascending=True, inplace=True)
        self.df_daily.drop_duplicates(subset='day', keep='first', inplace=True)
        self.save_dataframe_daily()
        print('[' + self.ticker + ']', (len(self.df_daily.index) - before), 'values added.')
        
    def update_intraday(self):
        self.load_dataframe_intraday()
        outputsize = 'full'
        if len(self.df.index) != 0 and ((datetime.datetime.now() - datetime.timedelta(minutes=570)) - self.df.date[-1]).seconds < self.interval * 600:
            outputsize = 'compact'
        req = requests.get('https://www.alphavantage.co/query?function=TIME_SERIES_INTRADAY&symbol=' + self.ticker + '&interval=' + str(self.interval) + 'min&outputsize=' + outputsize + '&apikey=' + self.api_key()).json()
        self.ticker = req['Meta Data']['2. Symbol']
        before = len(self.df.index)
        # Load new values
        new_df = pd.DataFrame.from_dict(req[[k for k, v in req.items()][-1]], orient="index").reset_index()
        new_df.columns = ['date', 'open', 'high', 'low', 'close', 'volume']
        new_df.date = pd.to_datetime(new_df.date, format='%Y-%m-%d %H:%M:%S', exact=False)
        new_df.open = pd.to_numeric(new_df.open)
        new_df.high = pd.to_numeric(new_df.high)
        new_df.low = pd.to_numeric(new_df.low)
        new_df.close = pd.to_numeric(new_df.close)
        new_df.volume = pd.to_numeric(new_df.volume)
        new_df.fillna(0, inplace=True)
        new_df.sort_values(by=['date'], kind='quicksort', ascending=True, inplace=True)
        # Merge dataframes
        self.df = new_df.append(self.df, ignore_index=True)
        self.df.date = pd.to_datetime(self.df.date, format='%Y-%m-%d %H:%M:%S', exact=False)
        self.df.open = pd.to_numeric(self.df.open)
        self.df.high = pd.to_numeric(self.df.high)
        self.df.low = pd.to_numeric(self.df.low)
        self.df.close = pd.to_numeric(self.df.close)
        self.df.volume = pd.to_numeric(self.df.volume)
        self.df.fillna(0, inplace=True)
        self.df.drop_duplicates(subset='date', keep='first', inplace=True)
        self.df.sort_values(by=['date'], kind='mergesort', ascending=True, inplace=True)
        self.df.drop_duplicates(subset='date', keep='first', inplace=True)
        self.save_dataframe_intraday()
        if self.ticker.startswith('NSE') or self.ticker.startswith('BSE'):
            self.df.date = self.df.date.dt.tz_localize(tz='US/Eastern').dt.tz_convert('Asia/Kolkata')
            print('[' + self.ticker + ']', 'Intraday :', str(self.df.date[-1:]).split('\n')[0].split(' ')[-2] + ' ' + str(self.df.date[-1:]).split('\n')[0].split(' ')[-1][:8])
        else:
            print('[' + self.ticker + ']', 'Intraday :', req['Meta Data']['3. Last Refreshed'])
        print('[' + self.ticker + ']', (len(self.df.index) - before), 'values added.')
